#!usr/bin/env python

import math
import h5py
import numpy as np
import sys
import os
import matplotlib.pyplot as plt
import openpyxl

# constants
error_allowed = 1.0E-6
lowest_energy_factor = 0.8  # *energy_grid.min()
boltzmann_constant = 8.617E-11  # MeV/K
title_font_size = 26
label_size = 20
label_figure_size = 20
line_width = 3.0
legend_font_size = 26
figure_size = [20, 16]
fissile = [92235, 94239]
none_fission = [92238]
New_Path = "Figure"

folder_sign = "\\"

# folder check
folder_num = 1
while True:
    try:
        test = open(New_Path + "/test.test", "a")
    except IOError:
        break
    else:
        test.close()
        os.system("del " + New_Path + folder_sign + "test.test")
        New_Path = "Figure" + str(folder_num)
        folder_num += 1
print 'Results will be saved in folder "' + New_Path + '"'


# Class Isotope, Cell, Inputdata are used to save the data from the hdf5 file
class Isotope:
    def __init__(self):
        self.zaid_id = 0
        self.num_density = 0.0E+00

        self.rr_abs = np.array([])
        self.rr_nufiss = np.array([])
        self.rr_scatter = np.array([])
        self.rr_detailed_scatter = np.array([])

        self.micro_abs_XS = np.array([])
        self.macro_abs_XS = np.array([])
        self.micro_nufiss_v_XS = np.array([])
        self.macro_nufiss_v_XS = np.array([])
        self.micro_scatter_XS = np.array([])
        self.macro_scatter_XS = np.array([])

        # the error used to calculate the difference of kinf
        self.error_xs_abs = np.array([])
        self.error_xs_nufiss = np.array([])
        self.error_rate_abs = np.array([])
        self.error_rate_nufiss = np.array([])

        self.kinf_xs_abs = np.array([])
        self.kinf_xs_nufiss = np.array([])
        self.kinf_rate_abs = np.array([])
        self.kinf_rate_nufiss = np.array([])

    def calculate_kinf_abs(self, fission, absorption):
        self.kinf_xs_abs = fission / (absorption + self.error_xs_abs)
        self.kinf_rate_abs = fission / (absorption + self.error_rate_abs)

    def calculate_kinf_nufiss(self, fission, absorption):
        self.kinf_xs_nufiss = (fission + self.error_xs_nufiss) / absorption
        self.kinf_rate_nufiss = (fission + self.error_rate_nufiss) / absorption


class Cell:
    def __init__(self):
        self.mat_id = 0
        self.isotopes = []
        self.scalar_flux = []
        self.volume = 0.0E+00

    # Because the following calculation use the flux data of the cell, so they are set in the Class Cell rather than
    # Class Isotopes
    def calculate_abs_xs(self):
        for index in range(len(self.isotopes)):
            self.isotopes[index].macro_abs_XS = self.isotopes[index].rr_abs / self.scalar_flux
            self.isotopes[index].micro_abs_XS = self.isotopes[index].macro_abs_XS / self.isotopes[index].num_density

    def calculate_nufiss_v_xs(self):
        for index in range(len(self.isotopes)):
            self.isotopes[index].macro_nufiss_v_XS = self.isotopes[index].rr_nufiss / self.scalar_flux
            self.isotopes[index].micro_nufiss_v_XS = self.isotopes[index].macro_nufiss_v_XS / self.isotopes[
                index].num_density

    def calculate_scatter_xs(self):
        for index in range(len(self.isotopes)):
            self.isotopes[index].macro_scatter_XS = self.isotopes[index].rr_scatter / self.scalar_flux
            self.isotopes[index].micro_scatter_XS = self.isotopes[index].macro_scatter_XS / self.isotopes[
                index].num_density

    def calculate_kinf_abs(self, fission, absorption):
        for index in range(len(self.isotopes)):
            self.isotopes[index].calculate_kinf_abs(fission, absorption)

    def calculate_kinf_nufiss(self, fission, absorption):
        for index in range(len(self.isotopes)):
            self.isotopes[index].calculate_kinf_nufiss(fission, absorption)


class Inputdata:
    def __init__(self):
        self.keff = 0.0E+00
        self.cell_number = 0
        self.max_isotope_number = 0
        self.energy_bins_number = 0
        self.energy_grid = np.array([0])
        self.cell_list = []
        self.material_sp = []
        # scatter_type
        # 0: scatter
        # 1: inscatter, outscatter, selfscatter
        self.scatter_type = 0

    def get_data_from_hdf5(self, filename):
        inputfile = h5py.File(filename, "r")
        state1 = inputfile["STATE_0001"]
        reaction_rate = state1["reaction_rate_details"]
        try:
            self.keff = state1["keff"][...][0]
        except IndexError:
            self.keff = state1["keff"][...]
        shape = reaction_rate["rr_abs"].shape
        self.max_isotope_number = shape[6]
        self.energy_bins_number = shape[5]
        self.cell_number = shape[4]

        rr_isoID = reaction_rate["rr_isoID"][...]
        rr_mID = reaction_rate["rr_mID"][...]
        rr_numDen = reaction_rate["rr_numDen"][...]
        self.energy_grid = reaction_rate["rr_eMesh"][...]
        rr_abs = reaction_rate["rr_abs"][...]
        rr_nufiss = reaction_rate["rr_nufiss"][...]
        rr_flux = reaction_rate["rr_flux"][...]
        volume = reaction_rate["rr_vol"][...]

        # get the normalization factor
        normalize_factor = normalize(volume, rr_flux)
        rr_abs /= normalize_factor
        rr_nufiss /= normalize_factor

        # just one scatter or three types
        try:
            rr_scatter = reaction_rate["rr_scatter"][...] / normalize_factor
            self.scatter_type = 0
        except KeyError:
            rr_scatter = (reaction_rate["rr_outscatter"][...] + reaction_rate["rr_selfscatter"][...]) / normalize_factor
            self.scatter_type = 1
            rr_detailed_scatter = np.array(([reaction_rate["rr_inscatter"][...], reaction_rate["rr_outscatter"][...],
                                             reaction_rate["rr_selfscatter"][...]])) / normalize_factor

        inputfile.close()

        # split the data into corresponging cell and then isotopes
        for cell_index in range(self.cell_number):
            self.cell_list.append(Cell())
            self.cell_list[cell_index].mat_id = rr_mID[0, 0, 0, 0, cell_index]
            self.cell_list[cell_index].scalar_flux = rr_flux[0, 0, 0, 0, cell_index, 0:self.energy_bins_number]
            self.cell_list[cell_index].volume = volume[0, 0, 0, 0, cell_index]
            for iso_index in range(self.max_isotope_number):
                iso = Isotope()
                iso.zaid_id = rr_isoID[0, 0, 0, 0, cell_index, iso_index]
                if iso.zaid_id == 0:
                    break
                iso.num_density = rr_numDen[0, 0, 0, 0, cell_index, iso_index]
                iso.rr_abs = rr_abs[0, 0, 0, 0, cell_index, 0:self.energy_bins_number, iso_index]
                iso.rr_nufiss = rr_nufiss[0, 0, 0, 0, cell_index, 0:self.energy_bins_number, iso_index]
                iso.rr_scatter = rr_scatter[0, 0, 0, 0, cell_index, 0:self.energy_bins_number, iso_index]
                try:
                    iso.rr_detailed_scatter = \
                        rr_detailed_scatter[0:3, 0, 0, 0, 0, cell_index, 0:self.energy_bins_number, iso_index]
                except UnboundLocalError:
                    pass
                self.cell_list[cell_index].isotopes.append(iso)
        self.material_specification()
        pass

    def material_specification(self):
        mat_index = [None] * self.cell_number
        for index in range(self.cell_number):
            mat_index[index] = self.cell_list[index].mat_id
        former_index = -1
        container = []
        for index in range(self.cell_number):
            if (mat_index[index] != former_index) and container:
                self.material_sp.append(container)
                container = [index]
                former_index = mat_index[container[0]]
            else:
                container.append(index)
                former_index = mat_index[index]
        if container:
            self.material_sp.append(container)
        for index1 in range(len(self.material_sp)):
            for index2 in self.material_sp[index1]:
                self.cell_list[index2].mat_id = index1 + 1
        self.calculate_abs_xs()
        self.calculate_nufiss_v_xs()
        self.calculate_scatter_xs()
        self.calculate_kinf()
        pass

    def calculate_abs_xs(self):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_abs_xs()

    def calculate_nufiss_v_xs(self):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_nufiss_v_xs()

    def calculate_scatter_xs(self):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_scatter_xs()

    # calculate kinf with the data of absorption and nuclear fission reaction rates
    def calculate_kinf(self):
        kinf = [0.0, 0.0, 0.0]
        for cell_ind in range(self.cell_number):
            produce = 0.0
            absorption = 0.0
            for e_ind in range(self.energy_bins_number):
                for iso_ind in range(len(self.cell_list[cell_ind].isotopes)):
                    produce += self.cell_list[cell_ind].isotopes[iso_ind].rr_nufiss[e_ind]
                    absorption += self.cell_list[cell_ind].isotopes[iso_ind].rr_abs[e_ind]
            kinf[0] += produce * self.cell_list[cell_ind].volume
            kinf[1] += absorption * self.cell_list[cell_ind].volume
        kinf[2] = kinf[0] / kinf[1]
        return kinf

    # the following two functions may have no use in this analysis
    def calculate_kinf_abs(self, reference):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_kinf_abs(reference[0], reference[1])

    def calculate_kinf_nufiss(self, reference):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_kinf_nufiss(reference[0], reference[1])


# self.scatter_type have not been defined
# this is aiming for the comparison of the two sets of values
class Reaction_rate_errors:
    def __init__(self):
        self.keff_reference = np.array([0.0, 0.0])
        self.cell_number = 0
        self.cell_list = []
        self.max_isotope_number = 0
        self.energy_bins_number = 0
        self.energy_grid = np.array([])
        self.kinf = np.array([0.0, 0.0, 0.0])
        # scatter_type
        # 0: scatter
        # 1: inscatter, outscatter, selfscatter
        self.scatter_type = 0
        self.figure_num = 1
        self.total_flux = np.array([])

    def get_data_from_inputdata(self, reference, testfile, option="m"):
        if option == "m":
            # the conditions of the following two "if"s are the check of this kind of reaction rates analysis
            if (len(reference.material_sp) == len(testfile.material_sp)) and (
                        reference.max_isotope_number == reference.max_isotope_number):
                self.cell_number = len(reference.material_sp)
                if np.linalg.norm(1.0 - testfile.energy_grid / reference.energy_grid) < error_allowed:
                    self.energy_grid = reference.energy_grid
                    self.keff_reference = np.array([reference.keff, testfile.keff])
                    self.max_isotope_number = reference.max_isotope_number
                    self.energy_bins_number = reference.energy_bins_number
                    self.kinf = reference.calculate_kinf()
                    # many values in different cells whose materials are the same have to be merged
                    for index in range(self.cell_number):
                        self.cell_list.append(Cell_errors())
                        self.cell_list[index].mat_id = reference.cell_list[reference.material_sp[index][0]].mat_id
                        # volume and scalar_flux
                        self.cell_list[index].scalar_flux = np.zeros((2, self.energy_bins_number))
                        for c_i in range(len(reference.material_sp[index])):
                            cell_no = reference.material_sp[index][c_i]
                            self.cell_list[index].volume[0] += reference.cell_list[cell_no].volume
                            self.cell_list[index].scalar_flux[0] += reference.cell_list[cell_no].scalar_flux * \
                                                                    reference.cell_list[cell_no].volume

                        for c_i in range(len(testfile.material_sp[index])):
                            cell_no = testfile.material_sp[index][c_i]
                            self.cell_list[index].volume[1] += testfile.cell_list[cell_no].volume
                            self.cell_list[index].scalar_flux[1] += testfile.cell_list[cell_no].scalar_flux * \
                                                                    testfile.cell_list[cell_no].volume
                        self.cell_list[index].scalar_flux[0] /= self.cell_list[index].volume[0]
                        self.cell_list[index].scalar_flux[1] /= self.cell_list[index].volume[1]
                        # isotopes
                        self.cell_list[index].isotope_num = len(
                            reference.cell_list[reference.material_sp[index][0]].isotopes)
                        for iso_ind in range(self.cell_list[index].isotope_num):
                            self.cell_list[index].isotopes.append(Isotope_errors())
                            self.cell_list[index].isotopes[iso_ind].zaid_id = \
                                reference.cell_list[reference.material_sp[index][0]].isotopes[iso_ind].zaid_id
                            self.cell_list[index].isotopes[iso_ind].num_density = \
                                reference.cell_list[reference.material_sp[index][0]].isotopes[0].num_density
                            self.cell_list[index].isotopes[iso_ind].rr_abs = np.zeros((2, self.energy_bins_number))
                            self.cell_list[index].isotopes[iso_ind].rr_scatter = np.zeros((2, self.energy_bins_number))
                            self.cell_list[index].isotopes[iso_ind].rr_nufiss = np.zeros((2, self.energy_bins_number))
                            self.cell_list[index].isotopes[iso_ind].rr_detailed_scatter = np.zeros(
                                (2, 3, self.energy_bins_number))
                            for c_i in range(len(reference.material_sp[index])):
                                cell_no = reference.material_sp[index][c_i]
                                self.cell_list[index].isotopes[iso_ind].rr_abs[0] += \
                                    reference.cell_list[cell_no].isotopes[iso_ind].rr_abs * \
                                    reference.cell_list[cell_no].volume
                                self.cell_list[index].isotopes[iso_ind].rr_nufiss[0] += \
                                    reference.cell_list[cell_no].isotopes[iso_ind].rr_nufiss * \
                                    reference.cell_list[cell_no].volume
                                self.cell_list[index].isotopes[iso_ind].rr_scatter[0] += \
                                    reference.cell_list[cell_no].isotopes[iso_ind].rr_scatter * \
                                    reference.cell_list[cell_no].volume
                                try:
                                    self.cell_list[index].isotopes[iso_ind].rr_detailed_scatter[0] += \
                                        (reference.cell_list[cell_no].isotopes[iso_ind].rr_detailed_scatter *
                                         reference.cell_list[cell_no].volume)
                                except ValueError:
                                    pass
                            self.cell_list[index].isotopes[iso_ind].rr_abs[0] /= self.cell_list[index].volume[0]
                            self.cell_list[index].isotopes[iso_ind].rr_nufiss[0] /= self.cell_list[index].volume[0]
                            self.cell_list[index].isotopes[iso_ind].rr_scatter[0] /= self.cell_list[index].volume[0]
                            self.cell_list[index].isotopes[iso_ind].rr_detailed_scatter[0] /= self.cell_list[
                                index].volume[0]
                            for c_i in range(len(testfile.material_sp[index])):
                                cell_no = testfile.material_sp[index][c_i]
                                self.cell_list[index].isotopes[iso_ind].rr_abs[1] += \
                                    testfile.cell_list[cell_no].isotopes[iso_ind].rr_abs * \
                                    testfile.cell_list[cell_no].volume
                                self.cell_list[index].isotopes[iso_ind].rr_nufiss[1] += \
                                    testfile.cell_list[cell_no].isotopes[iso_ind].rr_nufiss * \
                                    testfile.cell_list[cell_no].volume
                                self.cell_list[index].isotopes[iso_ind].rr_scatter[1] += \
                                    testfile.cell_list[cell_no].isotopes[iso_ind].rr_scatter * \
                                    testfile.cell_list[cell_no].volume
                                try:
                                    self.cell_list[index].isotopes[iso_ind].rr_detailed_scatter[1] += \
                                        testfile.cell_list[cell_no].isotopes[iso_ind].rr_detailed_scatter * \
                                        testfile.cell_list[cell_no].volume
                                except ValueError:
                                    pass
                            self.cell_list[index].isotopes[iso_ind].rr_abs[1] /= self.cell_list[index].volume[1]
                            self.cell_list[index].isotopes[iso_ind].rr_nufiss[1] /= self.cell_list[index].volume[1]
                            self.cell_list[index].isotopes[iso_ind].rr_scatter[1] /= self.cell_list[index].volume[1]
                            self.cell_list[index].isotopes[iso_ind].rr_detailed_scatter[1] /= self.cell_list[
                                index].volume[1]
                else:
                    print "Error! The two cases have different energy groups"
                    exit(-1)

            else:
                print "Error! The two cases have different structure"
                exit(-1)

    # calculate the kinf errors, the XS and errors in the Class Isotope are also calculated
    def calculate_errors_and_kinfs(self):
        self.calculate_kinf_abs()
        self.calculate_kinf_nufiss()
        self.calculate_scatter_xs()
        pass

    # plot the figures
    def plotting_results(self):
        self.plot_spectrum()
        self.plot_reactivity_dif()
        self.plot_relative_error()
        plt.close()

    def plot_spectrum(self):
        self.total_flux = np.zeros((2, self.energy_bins_number))
        spectrum = plt.figure(self.figure_num)
        self.figure_num += 1
        spectrum.set_size_inches(figure_size[0], figure_size[1] * 0.75)
        val_x1 = np.append(self.energy_grid, [lowest_energy_factor * self.energy_grid.min()])
        # volume multiplied
        for index in range(self.cell_number):
            self.total_flux[0] += self.cell_list[index].scalar_flux[0] * self.cell_list[index].volume[0]
        for index in range(self.cell_number):
            self.total_flux[1] += self.cell_list[index].scalar_flux[1] * self.cell_list[index].volume[1]
        temp_energy_rate = np.zeros(self.energy_bins_number + 1)
        for rate in range(self.energy_bins_number):
            temp_energy_rate[rate + 1] = val_x1[rate] / val_x1[rate + 1]
        temp_energy_rate[0] = temp_energy_rate[1]
        temp_energy_rate = np.log10(temp_energy_rate)
        val_y1 = np.append(self.total_flux[0][0], self.total_flux[0]) / temp_energy_rate
        val_y2 = np.append(self.total_flux[1][0], self.total_flux[1]) / temp_energy_rate
        plt.title("Normalized flux", fontsize=title_font_size)
        plt.xlabel("Energy(eV)", fontsize=label_size)
        plt.ylabel("Normalized flux", fontsize=label_size)
        plt.xlim(val_x1.min(), val_x1.max())
        pic = plt.plot(val_x1, val_y1, "k", val_x1, val_y2, "r")
        plt.setp(pic, drawstyle='steps', linewidth=line_width + 0.5)
        plt.xscale("log")
        plt.tick_params(axis='x', labelsize=label_figure_size)
        plt.tick_params(axis='y', labelsize=label_figure_size)
        plt.legend(pic, ("MCNP5", "MPACT"), fontsize=legend_font_size, loc="best", numpoints=1)
        plt.grid(True)
        plt.savefig(New_Path + "/spectrum", dpi=300, bbox_inches='tight')

    def plot_reactivity_dif(self):
        excel_wb = openpyxl.Workbook()
        ws = excel_wb.active
        ws.title = "Errors"
        ws["A1"] = "cell"
        ws["B1"] = "isotope"
        ws["C1"] = "err abs(pcm)"
        ws["D1"] = "err nufiss(pcm)"
        row_number = 2
        for c_ind in range(self.cell_number):
            # check whether the folder "c_ind" exists
            try:
                testfile = open(New_Path + "/" + str(c_ind) + "/test.txt", "a")
            except IOError:
                os.system("mkdir " + New_Path + folder_sign + str(c_ind))
            else:
                testfile.close()
                os.system("del " + New_Path + folder_sign + str(c_ind) + folder_sign + "test.txt")
            for iso_ind in range(self.cell_list[c_ind].isotope_num):
                fig1 = plt.figure(self.figure_num)
                self.figure_num += 1
                fig1.set_size_inches(figure_size[0], figure_size[1])
                isotope_zaid = self.cell_list[c_ind].isotopes[iso_ind].zaid_id / 10
                # picture 1.2 reference subplot
                plt.subplot(2, 1, 1)
                val_x1 = np.append(self.energy_grid, [lowest_energy_factor * self.energy_grid.min()])
                val_y11 = np.append(self.cell_list[c_ind].isotopes[iso_ind].micro_abs_XS[0][0],
                                    self.cell_list[c_ind].isotopes[iso_ind].micro_abs_XS[0])
                if np.array_equal(self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[0],
                                  np.zeros(self.energy_bins_number)):
                    sub11 = plt.plot(val_x1, val_y11, "k")
                else:
                    val_y12 = np.append(self.cell_list[c_ind].isotopes[iso_ind].micro_nufiss_v_XS[0][0],
                                        self.cell_list[c_ind].isotopes[iso_ind].micro_nufiss_v_XS[0])
                    sub11 = plt.plot(val_x1, val_y11, "k", val_x1, val_y12, "r")
                plt.setp(sub11, linewidth=line_width, drawstyle="steps")
                plt.xscale("log")
                try:
                    plt.yscale("log")
                except ValueError:
                    print "The values for " + str(
                        isotope_zaid) + " have 0, so the scale of y axis is changed to linear."
                    plt.yscale("linear")
                plt.title("Reference effective XS for " + str(isotope_zaid), fontsize=title_font_size)
                plt.xlabel("Energy(eV)", fontsize=label_size)
                plt.ylabel("Referece cross section (bam)", fontsize=label_size)
                plt.xlim(val_x1.min(), val_x1.max())
                plt.tick_params(axis='x', labelsize=label_figure_size)
                plt.tick_params(axis='y', labelsize=label_figure_size)
                plt.grid(True)
                plt.legend(sub11, ("absorption XS", "nu*fission XS"), fontsize=legend_font_size, loc="best",
                           numpoints=1)
                # picture 1.2 reactivity difference (pcm)
                plt.subplot(2, 1, 2)
                delta_kinf21 = (self.cell_list[c_ind].isotopes[iso_ind].kinf_xs_abs - self.kinf[2]) * 1.0E+05
                val_y21 = np.append(delta_kinf21[0], delta_kinf21)
                delta_kinf23 = (self.cell_list[c_ind].isotopes[iso_ind].kinf_rate_abs - self.kinf[2]) * 1.0E+05
                ws["A" + str(row_number)] = c_ind
                ws["B" + str(row_number)] = isotope_zaid
                ws["C" + str(row_number)] = delta_kinf23.sum()
                val_y23 = np.append(delta_kinf23[0], delta_kinf23)
                if np.array_equal(self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[0],
                                  np.zeros(self.energy_bins_number)):
                    ws["D" + str(row_number)] = 0.0
                    sub12 = plt.plot(val_x1, val_y21, "k", val_x1, val_y23, "k--")
                else:
                    delta_kinf22 = (self.cell_list[c_ind].isotopes[iso_ind].kinf_xs_nufiss - self.kinf[
                        2]) * 1.0E+05
                    val_y22 = np.append(delta_kinf22[0], delta_kinf22)
                    delta_kinf24 = (self.cell_list[c_ind].isotopes[iso_ind].kinf_rate_nufiss - self.kinf[
                        2]) * 1.0E+05
                    ws["D" + str(row_number)] = delta_kinf24.sum()
                    val_y24 = np.append(delta_kinf24[0], delta_kinf24)
                    sub12 = plt.plot(val_x1, val_y21, "k", val_x1, val_y22, "r",
                                     val_x1, val_y23, "k--", val_x1, val_y24, "r--")
                plt.setp(sub12, linewidth=line_width, drawstyle="steps")
                plt.xscale("log")
                plt.title("Error in reactivity compared with MCNP", fontsize=title_font_size)
                plt.xlabel("Energy(eV)", fontsize=label_size)
                plt.ylabel("Reactivity difference (pcm)", fontsize=label_size)
                plt.xlim(val_x1.min(), val_x1.max())
                plt.tick_params(axis='x', labelsize=label_figure_size)
                plt.tick_params(axis='y', labelsize=label_figure_size)
                plt.grid(True)
                if np.array_equal(self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[0],
                                  np.zeros(self.energy_bins_number)):
                    plt.legend(sub12, ("abs XS err", "abs rate err"), fontsize=legend_font_size, loc="best",
                               numpoints=1)
                else:
                    plt.legend(sub12, ("abs XS err", "nu*fis XS err", "abs rate err", "nu*fis rate err"),
                               fontsize=legend_font_size, loc="best", numpoints=1)
                plt.savefig(New_Path + "/" + str(c_ind) + "/Erg" + str(isotope_zaid) + "_pcm", dpi=300,
                            bbox_inches='tight')
                plt.close()
                row_number += 1
        ws["A" + str(row_number)] = "total"
        ws["C" + str(row_number)] = "=SUM(C2:C" + str(row_number - 1) + ")"
        ws["D" + str(row_number)] = "=SUM(D2:D" + str(row_number - 1) + ")"
        excel_wb.save(New_Path + "/errors.xlsx")

    def plot_relative_error(self):
        for c_ind in range(self.cell_number):
            # check whether the folder "c_ind" exists
            try:
                test = open(New_Path + "/" + str(c_ind) + "/test.txt", "a")
            except IOError:
                os.system("mkdir " + New_Path + folder_sign + str(c_ind))
            else:
                test.close()
                os.system("del " + New_Path + folder_sign + str(c_ind) + folder_sign + "test.txt")
            for iso_ind in range(self.cell_list[c_ind].isotope_num):
                fig2 = plt.figure(self.figure_num)
                self.figure_num += 1
                fig2.set_size_inches(figure_size[0], figure_size[1])
                isotope_zaid = self.cell_list[c_ind].isotopes[iso_ind].zaid_id / 10
                # picture 1.1 reference
                fig2.add_subplot(2, 1, 1)
                val_x1 = np.append(self.energy_grid, [lowest_energy_factor * self.energy_grid.min()])
                val_y11 = np.append(self.cell_list[c_ind].isotopes[iso_ind].micro_abs_XS[0][0],
                                    self.cell_list[c_ind].isotopes[iso_ind].micro_abs_XS[0])
                if np.array_equal(self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[0],
                                  np.zeros(self.energy_bins_number)):
                    sub21 = plt.plot(val_x1, val_y11, "k")
                else:
                    val_y12 = np.append(self.cell_list[c_ind].isotopes[iso_ind].micro_nufiss_v_XS[0][0],
                                        self.cell_list[c_ind].isotopes[iso_ind].micro_nufiss_v_XS[0])
                    sub21 = plt.plot(val_x1, val_y11, "k", val_x1, val_y12, "r")
                plt.setp(sub21, linewidth=line_width, drawstyle="steps")
                plt.xscale("log")
                try:
                    plt.yscale("log")
                except ValueError:
                    print "The values for " + str(
                        isotope_zaid) + " have 0, so the scale of y axis is changed to linear."
                    plt.yscale("linear")
                plt.title("Reference effective XS for " + str(isotope_zaid), fontsize=title_font_size)
                plt.xlabel("Energy(eV)", fontsize=label_size)
                plt.ylabel("Referece cross section (bam)", fontsize=label_size)
                plt.xlim(val_x1.min(), val_x1.max())
                plt.tick_params(axis='x', labelsize=label_figure_size)
                plt.tick_params(axis='y', labelsize=label_figure_size)
                plt.grid(True)
                plt.legend(sub21, ("absorption XS", "nu*fission XS"), fontsize=legend_font_size, loc="best",
                           numpoints=1)
                # picture 1.2 reactivity difference (pcm)
                fig2.add_subplot(212)
                delta_kinf21 = (self.cell_list[c_ind].isotopes[iso_ind].macro_abs_XS[1] /
                                self.cell_list[c_ind].isotopes[iso_ind].macro_abs_XS[0] - 1.0) * 100.0
                val_y21 = np.append(delta_kinf21[0], delta_kinf21)
                delta_kinf23 = (self.cell_list[c_ind].isotopes[iso_ind].macro_abs_XS[1] *
                                self.cell_list[c_ind].scalar_flux[1] /
                                self.cell_list[c_ind].isotopes[iso_ind].macro_abs_XS[0] /
                                self.cell_list[c_ind].scalar_flux[0] - 1.0) * 100.0
                val_y23 = np.append(delta_kinf23[0], delta_kinf23)
                if np.array_equal(self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[0],
                                  np.zeros(self.energy_bins_number)):
                    sub22 = plt.plot(val_x1, val_y21, "k",
                                     val_x1, val_y23, "k--")
                else:
                    delta_kinf22 = (self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[1] /
                                    self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[0] - 1.0) * 100.0
                    val_y22 = np.append(delta_kinf22[0], delta_kinf22)
                    delta_kinf24 = (self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[1] *
                                    self.cell_list[c_ind].scalar_flux[1] /
                                    self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[0] /
                                    self.cell_list[c_ind].scalar_flux[0] - 1.0) * 100.0
                    val_y24 = np.append(delta_kinf24[0], delta_kinf24)
                    sub22 = plt.plot(val_x1, val_y21, "k", val_x1, val_y22, "r",
                                     val_x1, val_y23, "k--", val_x1, val_y24, "r--")
                plt.setp(sub22, linewidth=line_width, drawstyle="steps")
                plt.xscale("log")
                plt.title("Relative error(%) compared with MCNP", fontsize=title_font_size)
                plt.xlabel("Energy(eV)", fontsize=label_size)
                plt.ylabel("Relative error (%)", fontsize=label_size)
                plt.xlim(val_x1.min(), val_x1.max())
                # plt.xlim(0.99 * val_x1.min(), 1.01 * val_x1.max())
                plt.tick_params(axis='x', labelsize=label_figure_size)
                plt.tick_params(axis='y', labelsize=label_figure_size)
                plt.grid(True)
                if np.array_equal(self.cell_list[c_ind].isotopes[iso_ind].macro_nufiss_v_XS[0],
                                  np.zeros(self.energy_bins_number)):
                    plt.legend(sub22, ("abs XS err", "abs rate err"),
                               fontsize=legend_font_size, loc="best",
                               numpoints=1)
                else:
                    plt.legend(sub22, ("abs XS err", "nu*fis XS err", "abs rate err",
                                       "nu*fis rate err"),
                               fontsize=legend_font_size, loc="best", numpoints=1)
                plt.savefig(New_Path + "/" + str(c_ind) + "/Erg" + str(isotope_zaid) + "_pct", dpi=300,
                            bbox_inches='tight')
                plt.close()

    # the following functions should just be called inside other functions, that is, they are better private
    def calculate_abs_xs(self):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_abs_xs()

    def calculate_nufiss_v_xs(self):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_nufiss_v_xs()

    def calculate_scatter_xs(self):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_scatter_xs()

    def calculate_kinf_abs(self):
        self.calculate_abs_xs()
        self.calculate_errors_abs()
        for index in range(self.cell_number):
            self.cell_list[index].calculate_kinf_abs(self.kinf[0], self.kinf[1])

    def calculate_kinf_nufiss(self):
        self.calculate_nufiss_v_xs()
        self.calculate_errors_nufiss()
        for index in range(self.cell_number):
            self.cell_list[index].calculate_kinf_nufiss(self.kinf[0], self.kinf[1])

    def calculate_errors_abs(self):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_errors_abs()

    def calculate_errors_nufiss(self):
        for index in range(self.cell_number):
            self.cell_list[index].calculate_errors_nufiss()


# the comparison of Cells from two hdf5 files, consist of the cell list of Class Reaction_rate_errors
class Cell_errors:
    def __init__(self):
        self.mat_id = 0
        self.isotope_num = 0
        self.isotopes = []
        self.scalar_flux = []
        self.volume = np.array([0.0, 0.0])

    def calculate_abs_xs(self):
        for index in range(len(self.isotopes)):
            self.isotopes[index].macro_abs_XS = self.isotopes[index].rr_abs / self.scalar_flux
            self.isotopes[index].micro_abs_XS = self.isotopes[index].macro_abs_XS / self.isotopes[index].num_density

    def calculate_nufiss_v_xs(self):
        for index in range(len(self.isotopes)):
            self.isotopes[index].macro_nufiss_v_XS = self.isotopes[index].rr_nufiss / self.scalar_flux
            self.isotopes[index].micro_nufiss_v_XS = self.isotopes[index].macro_nufiss_v_XS / self.isotopes[
                index].num_density

    def calculate_scatter_xs(self):
        for index in range(len(self.isotopes)):
            self.isotopes[index].macro_scatter_XS = self.isotopes[index].rr_scatter / self.scalar_flux
            self.isotopes[index].micro_scatter_XS = self.isotopes[index].macro_scatter_XS / self.isotopes[
                index].num_density

    def calculate_kinf_abs(self, fission, absorption):
        for index in range(len(self.isotopes)):
            self.isotopes[index].calculate_kinf_abs(fission, absorption)

    def calculate_kinf_nufiss(self, fission, absorption):
        for index in range(len(self.isotopes)):
            self.isotopes[index].calculate_kinf_nufiss(fission, absorption)

    def calculate_errors_abs(self):
        for index in range(len(self.isotopes)):
            self.isotopes[index].error_xs_abs = (self.isotopes[index].macro_abs_XS[1] - self.isotopes[
                index].macro_abs_XS[0]) * self.volume[0] * self.scalar_flux[0]
            self.isotopes[index].error_rate_abs = (self.isotopes[index].rr_abs[1] - self.isotopes[index].rr_abs[0]) * \
                                                  self.volume[0]

    def calculate_errors_nufiss(self):
        for index in range(len(self.isotopes)):
            self.isotopes[index].error_xs_nufiss = (self.isotopes[index].macro_nufiss_v_XS[1] - self.isotopes[
                index].macro_nufiss_v_XS[0]) * self.volume[0] * self.scalar_flux[0]
            self.isotopes[index].error_rate_nufiss = (self.isotopes[index].rr_nufiss[1] - self.isotopes[
                index].rr_nufiss[0]) * self.volume[0]


# the comparison of Isotopes from two hdf5 files, consist of the isotope list of Class Cell_errors
class Isotope_errors:
    def __init__(self):
        self.zaid_id = 0
        self.num_density = 0.0E+00

        # (2,e_n)
        self.rr_abs = np.array([])
        self.rr_scatter = np.array([])
        self.rr_nufiss = np.array([])
        self.rr_detailed_scatter = np.array([])

        # (2,e_n)
        self.micro_abs_XS = np.array([])
        self.macro_abs_XS = np.array([])
        self.micro_nufiss_v_XS = np.array([])
        self.macro_nufiss_v_XS = np.array([])
        self.micro_scatter_XS = np.array([])
        self.macro_scatter_XS = np.array([])

        # (e_n)
        self.error_xs_abs = np.array([])
        self.error_xs_nufiss = np.array([])
        self.error_rate_abs = np.array([])
        self.error_rate_nufiss = np.array([])

        # (e_n)
        self.kinf_xs_abs = np.array([])
        self.kinf_xs_nufiss = np.array([])
        self.kinf_rate_abs = np.array([])
        self.kinf_rate_nufiss = np.array([])

    def calculate_kinf_abs(self, fission, absorption):
        self.kinf_xs_abs = fission / (absorption + self.error_xs_abs)
        self.kinf_rate_abs = fission / (absorption + self.error_rate_abs)

    def calculate_kinf_nufiss(self, fission, absorption):
        self.kinf_xs_nufiss = (fission + self.error_xs_nufiss) / absorption
        self.kinf_rate_nufiss = (fission + self.error_rate_nufiss) / absorption


# return the normalization factor
def normalize(volume, scalar_flux):
    shape = scalar_flux.shape
    cells_num = shape[4]
    energy_num = shape[5]
    factor = 0.0E+00
    for c_n in range(cells_num):
        cell_flux = 0.0E+00
        for e_n in range(energy_num):
            cell_flux += scalar_flux[0, 0, 0, 0, c_n, e_n]
        factor += (cell_flux * volume[0, 0, 0, 0, c_n])
    scalar_flux /= factor
    return factor


# check whether the two sets of values indicate the same geometry structure
def check(reference, testfile, option):
    if option == "m":
        # return True
        if (len(reference.material_sp) == len(testfile.material_sp)) and (
                    reference.max_isotope_number == testfile.max_isotope_number):
            for index in range(len(reference.material_sp)):
                volume = [0.0, 0.0]
                for c_ind in reference.material_sp[index]:
                    volume[0] += reference.cell_list[c_ind].volume
                for c_ind in testfile.material_sp[index]:
                    volume[1] += testfile.cell_list[c_ind].volume
                if math.fabs(volume[0] - volume[1]) > error_allowed:
                    print "These two cases have different volume."
                    return False
                c_num = [reference.material_sp[index][0], testfile.material_sp[index][0]]
                if len(reference.cell_list[c_num[0]].isotopes) == len(testfile.cell_list[c_num[1]].isotopes):
                    iso_num = len(reference.cell_list[c_num[0]].isotopes)
                    for iso_ind in range(iso_num):
                        if (reference.cell_list[c_num[0]].isotopes[iso_ind].zaid_id !=
                                testfile.cell_list[c_num[1]].isotopes[iso_ind].zaid_id):
                            print "These two cases have different isotopes."
                            print "Reference: " + reference.cell_list[c_num[0]].isotopes[iso_ind].zaid_id
                            print "Test file: " + testfile.cell_list[c_num[1]].isotopes[iso_ind].zaid_id
                            return False
                        if math.fabs(reference.cell_list[c_num[0]].isotopes[iso_ind].num_density - testfile.cell_list[
                            c_num[1]].isotopes[iso_ind].num_density) > error_allowed:
                            print "These two cases have different isotope parameters."
                            return False
                    return True
                else:
                    print "These two cases have different isotopes."
                    return False
    else:
        # for other types of analysis
        return False


# reference and testfile are two objects of the class "Inputdata"
# main process
def reaction_rates_analysis(reference, testfile, option):
    if option == "m":
        if check(reference, testfile, option):
            os.system("mkdir " + New_Path)
            analysis = Reaction_rate_errors()
            analysis.get_data_from_inputdata(reference, testfile)
            analysis.calculate_errors_and_kinfs()
            analysis.plotting_results()


# main function
def compare_and_plot():
    number = len(sys.argv)
    option = "m"
    if number == 3:
        option = "m"
    elif number == 4:
        option = sys.argv[3]
    else:
        print "Parameters error!"
        exit()
    reference = Inputdata()
    reference.get_data_from_hdf5(sys.argv[1])
    testfile = Inputdata()
    testfile.get_data_from_hdf5(sys.argv[2])

    reaction_rates_analysis(reference, testfile, option)

    pass


# main function
compare_and_plot()
